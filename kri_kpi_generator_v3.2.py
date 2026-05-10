#!/usr/bin/env python3
"""
KRI/KPI Generator v3.2 - EPSS from CSV.GZ + Enterprise Reporting
Vulnerability Risk Management metrics generator with live EPSS feed and professional output

Fetches EPSS data from: https://epss.empiricalsecurity.com/epss_scores-current.csv.gz
Fetches CISA KEV data from multiple sources
Generates enterprise-grade Excel reports with executive summary
"""

import argparse
import logging
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import requests
import subprocess
import tempfile
import shutil
import json
import os
from datetime import datetime, timedelta
from pathlib import Path

# ============================================================================
# CONFIGURATION
# ============================================================================

EPSS_URL = "https://epss.empiricalsecurity.com/epss_scores-current.csv.gz"
KEV_SOURCES = [
    "https://www.cisa.gov/sites/default/files/feeds/known_exploited_vulnerabilities.json",
    "https://cisa.gov/sites/default/files/feeds/known_exploited_vulnerabilities.json",
    "https://raw.githubusercontent.com/cisagov/known_exploited_vulnerabilities/main/known_exploited_vulnerabilities.json"
]
TIMEOUT = 30

# Risk thresholds
CRITICAL_UNPATCHED_THRESHOLD = 10
RISK_SCORE_THRESHOLD = 200
SLA_WARNING = 80
SLA_CRITICAL = 95

# ============================================================================
# LOGGING SETUP
# ============================================================================

logging.basicConfig(
    level=logging.INFO,
    format='  %(message)s'
)
logger = logging.getLogger(__name__)

# ============================================================================
# KRI/KPI GENERATOR CLASS
# ============================================================================

class KRIKPIGenerator:
    """Generates KRI/KPI metrics for vulnerability management"""
    
    def __init__(self, asset_file, open_vuln_file, closed_vuln_file=None, output_dir=".", use_cache=True):
        """Initialize generator with input files"""
        self.asset_file = asset_file
        self.open_vuln_file = open_vuln_file
        self.closed_vuln_file = closed_vuln_file
        self.output_dir = output_dir
        self.use_cache = use_cache
        self.cache_dir = None
        self.metrics = {}
        self.kev_data = {}
        self.epss_data = {}
        
        logger.info("KRI/KPI GENERATOR v3.2 - EPSS from CSV.GZ + Enterprise Reporting")
        logger.info("Live Feed Integration")
    
    def setup_cache(self):
        """Create temporary cache directory"""
        self.cache_dir = tempfile.mkdtemp(prefix="kri_kpi_")
        logger.info(f"Cache dir: {self.cache_dir}")
    
    def fetch_epss_data(self):
        """Fetch EPSS scores from CSV.GZ using curl + gunzip"""
        logger.info("[2/6] Fetching EPSS data from CSV.GZ...")
        
        try:
            import subprocess
            
            # Create temp files
            gz_file = f"{self.cache_dir}/epss_scores.csv.gz"
            csv_file = f"{self.cache_dir}/epss_scores.csv"
            
            # Try curl method first
            logger.info(f"  Downloading: {EPSS_URL}...")
            curl_cmd = f"curl -s -L -o {gz_file} {EPSS_URL}"
            result = subprocess.run(curl_cmd, shell=True, capture_output=True, timeout=TIMEOUT)
            
            # Check if file was downloaded
            if result.returncode == 0 and os.path.exists(gz_file) and os.path.getsize(gz_file) > 0:
                # File downloaded successfully, decompress
                logger.info("  Decompressing with gunzip...")
                unzip_cmd = f"gunzip -c {gz_file} > {csv_file}"
                result = subprocess.run(unzip_cmd, shell=True, capture_output=True, timeout=TIMEOUT)
                
                if result.returncode == 0 and os.path.exists(csv_file) and os.path.getsize(csv_file) > 0:
                    # Parse CSV - skip first line (metadata), read actual header
                    logger.info("  Parsing CSV...")
                    df = pd.read_csv(csv_file, skiprows=1)
                    
                    # Normalize column names to lowercase
                    df.columns = df.columns.str.lower().str.strip()
                    
                    # Store as dict: CVE -> EPSS score
                    count = 0
                    for _, row in df.iterrows():
                        cve = str(row.get('cve', '')).strip()
                        epss_val = row.get('epss', None)
                        
                        # Try to convert EPSS to float
                        try:
                            epss = float(epss_val)
                        except (ValueError, TypeError):
                            continue
                        
                        # Validate CVE and EPSS range
                        if cve and pd.notna(epss) and 0 <= epss <= 1.0:
                            self.epss_data[cve] = epss
                            count += 1
                    
                    logger.info(f"✓ Loaded {count} EPSS scores from CSV.GZ")
                    
                    # Cleanup temp files
                    try:
                        if os.path.exists(gz_file):
                            os.remove(gz_file)
                        if os.path.exists(csv_file):
                            os.remove(csv_file)
                    except:
                        pass
                    
                    return True
            
            # Fallback: log warning and use scan data
            logger.warning("  EPSS CSV.GZ download/extraction failed")
            logger.info("  Will use scan data EPSS values instead")
            
            # Cleanup
            for f in [gz_file, csv_file]:
                if os.path.exists(f):
                    try:
                        os.remove(f)
                    except:
                        pass
            
            return False
            
        except Exception as e:
            logger.warning(f"  EPSS fetch failed: {str(e)}")
            logger.info("  Will use scan data EPSS values instead")
            return False
    
    def fetch_kev_data(self):
        """Fetch CISA KEV data with fallback sources"""
        logger.info("[1/6] Fetching CISA KEV data...")
        
        for i, url in enumerate(KEV_SOURCES, 1):
            try:
                logger.info(f"  Attempting: {Path(url).name}...")
                response = requests.get(url, timeout=TIMEOUT)
                response.raise_for_status()
                
                data = response.json()
                vulnerabilities = data.get('vulnerabilities', [])
                
                for vuln in vulnerabilities:
                    cve = vuln.get('cveID', '').strip()
                    if cve:
                        self.kev_data[cve] = vuln
                
                logger.info(f"✓ KEV Database: {len(self.kev_data)} entries loaded")
                return True
                
            except Exception as e:
                logger.warning(f"    Failed: {str(e)[:80]}")
                continue
        
        logger.warning("⚠ KEV fetch from all sources failed")
        return False
    
    def load_data(self):
        """Load asset and vulnerability data"""
        logger.info("[3/6] Loading asset and vulnerability data...")
        
        # Load assets
        self.assets = pd.read_excel(self.asset_file)
        logger.info(f"  ✓ Loaded {len(self.assets)} assets")
        
        # Load open findings
        self.open_findings = pd.read_excel(self.open_vuln_file)
        self.open_findings['Discovery Date'] = pd.to_datetime(self.open_findings['Discovery Date'])
        logger.info(f"  ✓ Loaded {len(self.open_findings)} OPEN vulnerability records")
        
        # Load closed findings if provided
        if self.closed_vuln_file:
            self.closed_findings = pd.read_excel(self.closed_vuln_file)
            self.closed_findings['Discovery Date'] = pd.to_datetime(self.closed_findings['Discovery Date'])
            self.closed_findings['Remediation Date'] = pd.to_datetime(self.closed_findings['Remediation Date'])
            logger.info(f"  ✓ Loaded {len(self.closed_findings)} CLOSED vulnerability records")
        else:
            self.closed_findings = pd.DataFrame()
    
    def enrich_findings_with_epss(self):
        """Add fetched EPSS and KEV data to findings - optimized"""
        for df in [self.open_findings, self.closed_findings]:
            if df.empty:
                continue
            
            # Map from EPSS database first
            df['EPSS Final'] = df['CVE'].map(self.epss_data)
            
            # Fall back to scan data EPSS
            df['EPSS Final'] = df['EPSS Final'].fillna(df.get('EPSS', 0))
            
            # Default to 0 if missing
            df['EPSS Final'] = df['EPSS Final'].fillna(0).astype(float)
            
            # Mark KEV status
            df['Is KEV'] = df['CVE'].isin(self.kev_data.keys())
    
    def calculate_kpis(self):
        """Calculate 8 Key Performance Indicators"""
        logger.info("[4/6] Calculating KPIs...")
        
        # Merge with asset criticality
        merged = self.open_findings.merge(
            self.assets[['Asset ID', 'Business Criticality']], 
            on='Asset ID', 
            how='left'
        )
        merged['Criticality Weight'] = merged['Business Criticality'].map({
            'Tier 1': 3.0, 'Tier 2': 2.0, 'Tier 3': 1.0
        }).fillna(1.0)
        
        # KPI 1: MTTR by Severity
        if len(self.closed_findings) > 0:
            mttr_data = self.closed_findings.groupby('Severity').apply(
                lambda x: ((x['Remediation Date'] - x['Discovery Date']).dt.days.mean())
            ).to_dict()
            self.metrics['MTTR Critical'] = round(mttr_data.get('Critical', 0), 1)
            self.metrics['MTTR High'] = round(mttr_data.get('High', 0), 1)
            self.metrics['MTTR Medium'] = round(mttr_data.get('Medium', 0), 1)
            self.metrics['MTTR Low'] = round(mttr_data.get('Low', 0), 1)
            self.metrics['MTTR Info'] = round(mttr_data.get('Informational', 0), 1)
        else:
            self.metrics['MTTR Critical'] = 0.0
            self.metrics['MTTR High'] = 0.0
            self.metrics['MTTR Medium'] = 0.0
            self.metrics['MTTR Low'] = 0.0
            self.metrics['MTTR Info'] = 0.0
        
        # KPI 2: SLA Compliance - consolidated
        sla_targets = {'Critical': 14, 'High': 30, 'Medium': 60, 'Low': 120, 'Informational': 365}
        sla_values = {}
        sla_compliances = []
        
        for severity, target in sla_targets.items():
            severity_findings = self.open_findings[self.open_findings['Severity'] == severity]
            if len(severity_findings) > 0:
                compliant = sum((datetime.now() - severity_findings['Discovery Date']).dt.days <= target)
                compliance = (compliant / len(severity_findings)) * 100
            else:
                compliance = 100
            sla_values[severity] = round(compliance, 1)
            sla_compliances.append(compliance)
        
        # Store individual and overall
        for severity, compliance in sla_values.items():
            self.metrics[f'SLA Compliance {severity}'] = compliance
        
        overall_sla = sum(sla_compliances) / len(sla_compliances) if sla_compliances else 100
        self.metrics['SLA Compliance Overall %'] = round(overall_sla, 1)
        
        # KPI 3: Unpatched Critical/High Count
        self.metrics['Unpatched Critical'] = len(self.open_findings[self.open_findings['Severity'] == 'Critical'])
        self.metrics['Unpatched High'] = len(self.open_findings[self.open_findings['Severity'] == 'High'])
        
        # KPI 4: EPSS Distribution
        self.metrics['EPSS 0.0-0.25'] = len(self.open_findings[self.open_findings['EPSS Final'] < 0.25])
        self.metrics['EPSS 0.25-0.5'] = len(self.open_findings[(self.open_findings['EPSS Final'] >= 0.25) & (self.open_findings['EPSS Final'] < 0.5)])
        self.metrics['EPSS 0.5-0.75'] = len(self.open_findings[(self.open_findings['EPSS Final'] >= 0.5) & (self.open_findings['EPSS Final'] < 0.75)])
        self.metrics['EPSS 0.75+'] = len(self.open_findings[self.open_findings['EPSS Final'] >= 0.75])
        
        # KPI 5: Monthly Discovery Rate
        month_ago = datetime.now() - timedelta(days=30)
        recent = len(self.open_findings[self.open_findings['Discovery Date'] > month_ago])
        self.metrics['Monthly Discovery Rate'] = recent
        
        # KPI 6: Patch Lag (CISA KEV)
        kev_open = self.open_findings[self.open_findings['Is KEV']]
        if len(kev_open) > 0:
            lag = (datetime.now() - kev_open['Discovery Date']).dt.days.max()
            self.metrics['Patch Lag (KEV) Days'] = int(lag)
        else:
            self.metrics['Patch Lag (KEV) Days'] = 0
        
        # KPI 7: False Positive Rate
        fp_findings = self.open_findings[self.open_findings['Description'].str.contains('false', case=False, na=False)]
        self.metrics['False Positive Rate %'] = round(len(fp_findings) / len(self.open_findings) * 100 if len(self.open_findings) > 0 else 0, 1)
        
        # KPI 8: Detection Coverage by Scan Type
        if 'Scanner' in self.open_findings.columns:
            self.metrics['Detection Coverage'] = self.open_findings['Scanner'].value_counts().to_dict()
        else:
            self.metrics['Detection Coverage'] = {}
        
        logger.info(f"  ✓ Calculated 8 KPIs")
        logger.info(f"    • KEV Database: {len(self.kev_data)} entries loaded")
        logger.info(f"    • EPSS Database: {len(self.epss_data)} entries loaded")
    
    def calculate_kris(self):
        """Calculate 7 Key Risk Indicators"""
        logger.info("[5/6] Calculating KRIs...")
        
        # Merge open findings with asset criticality
        merged = self.open_findings.merge(
            self.assets[['Asset ID', 'Business Criticality']], 
            on='Asset ID', 
            how='left'
        )
        merged['Criticality Weight'] = merged['Business Criticality'].map({
            'Tier 1': 3.0, 'Tier 2': 2.0, 'Tier 3': 1.0
        }).fillna(1.0)
        
        # KRI 1: Unmitigated Critical Risk Score
        critical = merged[merged['Severity'] == 'Critical']
        risk_score = (critical['EPSS Final'] * critical['Criticality Weight']).sum() if len(critical) > 0 else 0
        self.metrics['Unmitigated Critical Risk Score'] = round(risk_score, 2)
        
        # KRI 2: Days at Risk (DAR)
        merged['Days at Risk'] = (datetime.now() - merged['Discovery Date']).dt.days
        merged['Risk Units'] = merged['EPSS Final'] * merged['Criticality Weight'] * merged['Days at Risk']
        self.metrics['Total Days at Risk'] = round(merged['Risk Units'].sum(), 1)
        
        # KRI 3: Exploit-in-the-Wild (KEV) Count
        kev_findings = self.open_findings[self.open_findings['Is KEV']]
        self.metrics['Exploit-in-the-Wild Count'] = len(kev_findings)
        
        # KRI 4: Control Effectiveness Gap
        mitigated = self.open_findings[self.open_findings['Description'].str.contains('mitigat', case=False, na=False)]
        gap = (len(mitigated) / len(self.open_findings) * 100) if len(self.open_findings) > 0 else 0
        self.metrics['Control Effectiveness Gap %'] = round(gap, 1)
        
        # KRI 5: Regression Rate (properly calculated)
        regressions = 0
        if len(self.closed_findings) > 0:
            closed_cves = set(self.closed_findings['CVE'].unique())
            open_cves = set(self.open_findings['CVE'].unique())
            regressed = closed_cves.intersection(open_cves)
            regressions = len(regressed)
            regression_rate = (regressions / len(self.closed_findings) * 100)
        else:
            regression_rate = 0
        self.metrics['Regression Rate %'] = round(regression_rate, 1)
        
        # KRI 6: Asset Risk Concentration
        top_10_assets = merged.groupby('Asset ID')['EPSS Final'].sum().nlargest(int(len(merged) * 0.1) or 1).sum()
        total_risk = merged['EPSS Final'].sum()
        concentration = (top_10_assets / total_risk * 100) if total_risk > 0 else 0
        self.metrics['Asset Risk Concentration %'] = round(concentration, 1)
        
        # KRI 7: Technical Debt (findings >120 days old)
        old_findings = merged[merged['Days at Risk'] > 120]
        self.metrics['Technical Debt Count'] = len(old_findings)
        
        logger.info(f"  ✓ Calculated 7 KRIs")
    
    def generate_excel_report(self):
        """Generate professional Excel report with multiple sheets"""
        logger.info("[6/6] Generating Excel metric deck...")
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"{self.output_dir}/KRI_KPI_Metrics_{timestamp}.xlsx"
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Metrics Summary"
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        title_font = Font(bold=True, size=14)
        alt_fill = PatternFill(start_color="E8EEF5", end_color="E8EEF5", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws['A1'] = "KRI/KPI Metrics Dashboard"
        ws['A1'].font = title_font
        ws.merge_cells('A1:B1')
        
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(size=10, italic=True)
        
        # Headers
        row = 4
        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Value"
        ws[f'C{row}'] = "Status"
        
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].fill = header_fill
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'{col}{row}'].border = border
        
        row += 1
        
        # Write metrics with status indicators
        for metric, value in self.metrics.items():
            if isinstance(value, dict):
                continue  # Skip dicts, handle separately
            
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            
            # Status indicator
            status = self._get_metric_status(metric, value)
            ws[f'C{row}'] = status
            
            # Styling
            if row % 2 == 0:
                ws[f'A{row}'].fill = alt_fill
                ws[f'B{row}'].fill = alt_fill
                ws[f'C{row}'].fill = alt_fill
            
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].border = border
            
            ws[f'B{row}'].alignment = Alignment(horizontal='right')
            
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        
        # Severity Breakdown sheet
        ws2 = wb.create_sheet("Severity Breakdown")
        severities = ['Critical', 'High', 'Medium', 'Low', 'Informational']
        
        ws2['A1'] = "Severity"
        ws2['B1'] = "Count"
        ws2['C1'] = "MTTR (days)"
        ws2['D1'] = "SLA %"
        
        for col in ['A', 'B', 'C', 'D']:
            ws2[f'{col}1'].fill = header_fill
            ws2[f'{col}1'].font = header_font
            ws2[f'{col}1'].border = border
        
        for i, severity in enumerate(severities, 2):
            count = len(self.open_findings[self.open_findings['Severity'] == severity])
            mttr = self.metrics.get(f'MTTR {severity}', 0)
            sla = self.metrics.get(f'SLA Compliance {severity}', 100)
            
            ws2[f'A{i}'] = severity
            ws2[f'B{i}'] = count
            ws2[f'C{i}'] = mttr
            ws2[f'D{i}'] = sla
            
            for col in ['A', 'B', 'C', 'D']:
                ws2[f'{col}{i}'].border = border
                if i % 2 == 0:
                    ws2[f'{col}{i}'].fill = alt_fill
        
        ws2.column_dimensions['A'].width = 15
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 15
        ws2.column_dimensions['D'].width = 12
        
        # KEV Analysis sheet
        ws3 = wb.create_sheet("KEV Analysis")
        kev_open = self.open_findings[self.open_findings['Is KEV']]
        
        ws3['A1'] = "CISA Known Exploited Vulnerabilities Analysis"
        ws3['A1'].font = title_font
        
        ws3['A3'] = "Metric"
        ws3['B3'] = "Value"
        ws3['A3'].fill = header_fill
        ws3['B3'].fill = header_fill
        ws3['A3'].font = header_font
        ws3['B3'].font = header_font
        
        metrics_kev = [
            ("Total KEV Entries in Database", len(self.kev_data)),
            ("Open KEV Findings", len(kev_open)),
            ("Coverage Gap %", round(100 - (len(kev_open) / len(self.open_findings) * 100), 1) if len(self.open_findings) > 0 else 0),
            ("Exploit Lag (days)", self.metrics.get('Patch Lag (KEV) Days', 0))
        ]
        
        row = 4
        for label, val in metrics_kev:
            ws3[f'A{row}'] = label
            ws3[f'B{row}'] = val
            row += 1
        
        # Detection Coverage sheet
        ws4 = wb.create_sheet("Coverage Analysis")
        if 'Scanner' in self.open_findings.columns:
            coverage = self.open_findings['Scanner'].value_counts()
            
            ws4['A1'] = "Detection Coverage by Scanner"
            ws4['A1'].font = title_font
            
            ws4['A3'] = "Scanner"
            ws4['B3'] = "Count"
            ws4['C3'] = "Coverage %"
            
            for col in ['A', 'B', 'C']:
                ws4[f'{col}3'].fill = header_fill
                ws4[f'{col}3'].font = header_font
                ws4[f'{col}3'].border = border
            
            row = 4
            for scanner, count in coverage.items():
                pct = (count / len(self.open_findings) * 100)
                ws4[f'A{row}'] = scanner
                ws4[f'B{row}'] = count
                ws4[f'C{row}'] = round(pct, 1)
                row += 1
            
            ws4.column_dimensions['A'].width = 20
            ws4.column_dimensions['B'].width = 12
            ws4.column_dimensions['C'].width = 15
        
        wb.save(output_file)
        logger.info(f"  ✓ Saved to {output_file}")
        
        return output_file
    
    def _get_metric_status(self, metric, value):
        """Determine status indicator (OK, WARN, CRITICAL)"""
        if 'Unpatched Critical' in metric and value > CRITICAL_UNPATCHED_THRESHOLD:
            return "🔴 CRIT"
        elif 'Risk Score' in metric and value > RISK_SCORE_THRESHOLD:
            return "🔴 CRIT"
        elif 'SLA' in metric and value < SLA_WARNING:
            return "🔴 CRIT"
        elif 'SLA' in metric and value < SLA_CRITICAL:
            return "🟡 WARN"
        elif 'KEV' in metric and value > 5:
            return "🟡 WARN"
        else:
            return "🟢 OK"
    
    def cleanup(self):
        """Archive cache"""
        if self.cache_dir and os.path.exists(self.cache_dir):
            archive_name = f"{self.output_dir}/kri_cache_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            shutil.move(self.cache_dir, archive_name)
    
    def run(self):
        """Execute full pipeline"""
        print()
        print("=" * 70)
        print("KRI/KPI GENERATOR v3.2 - EPSS from CSV.GZ + Enterprise Reporting")
        print("=" * 70)
        
        self.setup_cache()
        print()
        
        # Fetch data
        self.fetch_kev_data()
        self.fetch_epss_data()
        
        # Load and process
        self.load_data()
        print()
        
        self.enrich_findings_with_epss()
        
        # Calculate metrics
        self.calculate_kpis()
        self.calculate_kris()
        print()
        
        # Generate report
        output_file = self.generate_excel_report()
        
        # Cleanup
        self.cleanup()
        
        print()
        print("=" * 70)
        print("✓ COMPLETE - Live Threat Intelligence Integrated")
        print("=" * 70)
        print()
        
        return output_file

# ============================================================================
# MAIN
# ============================================================================

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="KRI/KPI Generator v3.2 - Vulnerability Risk Management Metrics",
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument('--assets', required=True, help='Asset inventory Excel file')
    parser.add_argument('--open', required=True, help='Open findings Excel file')
    parser.add_argument('--closed', help='Closed findings Excel file (optional)')
    parser.add_argument('--output', default='.', help='Output directory')
    parser.add_argument('--no-cache', action='store_true', help='Disable caching')
    
    args = parser.parse_args()
    
    generator = KRIKPIGenerator(
        asset_file=args.assets,
        open_vuln_file=getattr(args, 'open'),
        closed_vuln_file=args.closed,
        output_dir=args.output,
        use_cache=not args.no_cache
    )
    
    generator.run()
